import csv
import io
from pathlib import Path
from zipfile import ZipFile
import pytest
from openpyxl import load_workbook
from pypdf import PdfReader

RES_DIR = Path(__file__).parent/"resources"
ZIP_PATH = RES_DIR/"bundle.zip"

PDF_NAME = "example.pdf"
XLSX_NAME = "example.xlsx"
CSV_NAME = "example.csv"


def create_zip(): # создание зип
    if ZIP_PATH.exists():
        ZIP_PATH.unlink()
    with ZipFile(ZIP_PATH, mode="w") as z:
        z.write(RES_DIR / PDF_NAME, arcname=PDF_NAME)
        z.write(RES_DIR / XLSX_NAME, arcname=XLSX_NAME)
        z.write(RES_DIR / CSV_NAME, arcname=CSV_NAME)
    return ZIP_PATH

@pytest.fixture(scope="module")
def bundle_zip():
    return create_zip()


def test_read_pdf(bundle_zip): # проверка пдф
    with ZipFile(bundle_zip) as z:
        pdf_bytes = z.read(PDF_NAME)
    reader = PdfReader(io.BytesIO(pdf_bytes))
    text = "".join(page.extract_text() or "" for page in reader.pages)
    print(text)
    assert "Количество получателей ЖПС по годам согласно заявке" in text and "Февраль 1 0" in text




def test_read_xlsx(bundle_zip): # проверка xlsx
    with ZipFile(bundle_zip) as z:
        xlsx_bytes = z.read(XLSX_NAME)
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        assert ws.cell(row=4, column=1).value == "Вид по номенклатурной классификации медицинских изделий" and \
               ws.cell(row=4, column=2).value == "Наименование вида" and \
               ws.cell(row=4, column=3).value == "Средневзвешенная цена"

def test_read_csv(bundle_zip): # проверка цсв
    with ZipFile(bundle_zip) as z:
        with z.open(CSV_NAME) as raw:
            text_io = io.TextIOWrapper(raw, encoding="utf-8-sig")
            reader = csv.reader(text_io)
            rows = list(reader)
            for row in rows:
                print(row)
            assert ["1", "Аппарат УЗИ", "2", "1500000"] in rows
